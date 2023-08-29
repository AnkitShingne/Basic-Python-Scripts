import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb['Sheet']
pl = open('clubs.txt')
sk = open('sk.txt')
fis = []
fis.append(pl)
fis.append(sk)
i =0 
for file in fis:
	S = file.readlines()
	for i in range(1,len(S)+1):
		sheet[get_column_letter(fis.index(file) + 1) + str(i)] = S[i-1]
wb.save('text3.xlsx')
pl.close()
sk.close()
	
