import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font

wb = openpyxl.load_workbook('text3.xlsx')
sheet = wb['Sheet']
fis = []
for col in range(1,sheet.max_column + 1):
	for i in range(1,sheet.max_row + 1):
		fis.append(sheet[get_column_letter(col) + str(i)].value)

new = open('PL.txt','w')
random = open('random.txt','w')

for j in range(0,3):
    new.write(str(fis[j]))

for k in range(3,6):
	random.write(str(fis[k]))

new.close()
random.close()
	