#! python3
# multiplication.py - Creates table using commands from command line.

import openpyxl, sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font
wb = Workbook()
sheet = wb.active
sheet['A1'] = ''
for i in range(1,(int(sys.argv[1]) + 1)):
	sheet.cell(row=1, column=i+1).value = i
for j in range(1,(int(sys.argv[1])+1)):
	sheet.cell(row=j+1, column=1).value = j

for i in range(1,(int(sys.argv[1])+1)):
	for j in range(1,(int(sys.argv[1])+1)):
		sheet.cell(row=i+1, column=j+1).value = '=PRODUCT((' + str(sheet.cell(row=i+1, column=j).value) +'),(' + str(sheet.cell(row=i, column=j+1).value) +'))'
wb.save('table.xlsx')
	
