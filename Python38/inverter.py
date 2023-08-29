#! python3
# inverter.py - Inverts cell values.

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
row1 = []
for i in range(1, sheet.max_row + 1):
	fis = sheet.cell(i, 1).value
	row1.append(fis)
print(row1)
row2 = []
for j in range(1,sheet.max_row + 1):
	fus = sheet.cell(j, 2).value
	row2.append(fus)
print(row2)
for row in sheet['A1:B7']:
  for cell in row:
    cell.value = None
for k in range(1,8):
	sheet[get_column_letter(k) + '1'] = row1[k-1]
	sheet[get_column_letter(k) + '2'] = row2[k-1]

wb.save('rock4.xlsx')

